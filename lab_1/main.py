from schuhart import SchuhartControlCard
from openpyxl import load_workbook
import pandas as pd
import warnings
warnings.filterwarnings("ignore")


def main():
    wb = load_workbook('./lab_1.xlsx')
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        print(f'Таблица {sheet_name}')
        sheet = wb[sheet_name]
        data = sheet.values
        cols = next(data)
        df = pd.DataFrame(data, columns=cols)
        if None in cols:
            df = df.drop([None], axis=1)
        df = df.dropna()
        try:
            df['Дата'] = pd.to_datetime(df['Дата'])
        except Exception:
            df = df.rename(columns={'Сумма': 'Дата', 'Дата': 'Сумма'})
            df['Дата'] = pd.to_datetime(df['Дата'])
        finally:
            df['Сумма'] = [str(i).replace(' ', '')
                           .replace(',', '.') for i in df['Сумма']]
            df['Сумма'] = df['Сумма'].astype(float)
        card = SchuhartControlCard(df)
        card.calculate_card()
        print()
        card.draw_card(sheet_name)


if __name__ == "__main__":
    main()
